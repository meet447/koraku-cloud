"""Build .xlsx files from a JSON spec (CLI: python -m koraku.artifacts.xlsx_build)."""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any


def _load_spec(path: str | None, inline: str | None) -> dict[str, Any]:
    if path:
        return json.loads(Path(path).read_text(encoding="utf-8"))
    if inline:
        return json.loads(inline)
    return {}


def build_xlsx(spec: dict[str, Any], out_path: str) -> str:
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise SystemExit(
            "openpyxl is required. Install with: pip install 'koraku[artifacts]'"
        ) from e

    wb = Workbook()
    sheets = spec.get("sheets")
    if not sheets:
        sheets = [
            {
                "name": str(spec.get("sheet_name") or "Sheet1"),
                "headers": spec.get("headers") or [],
                "rows": spec.get("rows") or [],
            }
        ]

    first = True
    for sheet_spec in sheets:
        if not isinstance(sheet_spec, dict):
            continue
        if first:
            ws = wb.active
            ws.title = str(sheet_spec.get("name") or "Sheet1")[:31]
            first = False
        else:
            ws = wb.create_sheet(title=str(sheet_spec.get("name") or "Sheet")[:31])

        headers = sheet_spec.get("headers") or []
        if headers:
            for ci, header in enumerate(headers, start=1):
                ws.cell(row=1, column=ci, value=str(header))

        start_row = 2 if headers else 1
        for ri, row in enumerate(sheet_spec.get("rows") or [], start=start_row):
            if not isinstance(row, list):
                ws.cell(row=ri, column=1, value=str(row))
                continue
            for ci, val in enumerate(row, start=1):
                ws.cell(row=ri, column=ci, value=val)

        for col_idx in range(1, (len(headers) or 1) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 16

    out = Path(out_path).expanduser().resolve()
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    return str(out)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Build a .xlsx from JSON spec")
    parser.add_argument("--out", required=True, help="Output .xlsx path")
    parser.add_argument("--spec", help="Path to JSON spec file")
    parser.add_argument("--json", dest="inline_json", help="Inline JSON spec")
    args = parser.parse_args(argv)

    spec = _load_spec(args.spec, args.inline_json)
    path = build_xlsx(spec, args.out)
    print(json.dumps({"ok": True, "path": path, "type": "xlsx"}))
    return 0


if __name__ == "__main__":
    sys.exit(main())
